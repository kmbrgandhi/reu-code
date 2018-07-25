
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.patches as mpatches
import math
from conv_utils import con_psl_matrix

# Learn about API authentication here: https://plot.ly/python/getting-started
# Find your api_key here: https://plot.ly/settings/api


from openpyxl import load_workbook


def naive_time_plot():
    wb = load_workbook(filename = 'naive_exhaustive.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    lengths = []
    time_values = []
    for i in range(2, 23):
        lengths.append(sheet['D' + str(i)].value)
        time_values.append(sheet['G' + str(i)].value)

    fig = plt.figure()

    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.set_title('Naive Exhaustive Algorithm: Runtime vs. Length')

    ax.set_xlabel('length')
    ax.set_ylabel('time')

    ax.text(10, 400, 'exponential runtime', style='italic',
            bbox={'facecolor': 'red', 'alpha': 0.5, 'pad': 10})


    ax.scatter(lengths, time_values, s = 20)

    ax.axis([0, 25, 0, 900])

    plt.show()





def efficient_time_plot():
    wb = load_workbook(filename='efficient_exhaustive_two.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    lengths = []
    time_values = []
    for i in range(2, 27):
        lengths.append(sheet['C' + str(i)].value)
        time_values.append(sheet['E' + str(i)].value)

    fig = plt.figure()

    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.set_title('Efficient Exhaustive Algorithm: Runtime vs. Length')

    ax.set_xlabel('length')
    ax.set_ylabel('time')

    x_2 = np.linspace(2, 26, 200)
    y_2 = np.exp2(x_2) * time_values[0] * 0.25


    ax.text(5.5, 6, 'sub-exponential runtime', style='italic',
            bbox={'facecolor': 'red', 'alpha': 0.5, 'pad': 10})

    ax.scatter(lengths, time_values)
    ax.plot(x_2, y_2)

    ax.axis([0, 27, 0, 10])

    plt.show()


def genetic_parentselections_plot():
    wb = load_workbook(filename='genetic_half_crossover_test_binary.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    lengths = []
    part_random = []
    random = []
    tournament = []
    lengths.append(sheet['A' + str(4)].value)
    part_random.append(sheet['B' + str(4)].value)
    random.append(sheet['C' + str(4)].value)
    tournament.append(sheet['D' + str(4)].value)
    for i in range(5, 57):
        lengths.append(sheet['A' + str(i)].value)
        part_random.append(sheet['B' + str(i)].value + part_random[i-5])
        random.append(sheet['C' + str(i)].value + random[i-5])
        tournament.append(sheet['D' + str(i)].value+ tournament[i-5])

    fig = plt.figure()

    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.set_title('Comparison of Parent Selection Variants')

    ax.set_xlabel('length')
    ax.set_ylabel('cumulative distance from optimal peak sidelobe value')

    ax.scatter(lengths, random, c='r')
    ax.plot(lengths, random, c='r')
    ax.scatter(lengths, part_random, c='b')
    ax.plot(lengths, part_random, c='b')
    ax.scatter(lengths, tournament, c='g')
    ax.plot(lengths, tournament, c='g')

    classes = ['Tournament', 'Random', 'Partially Random']
    recs = [mpatches.Rectangle((0, 0), 1, 1, fc='g'), mpatches.Rectangle((0, 0), 1, 1, fc='r'),
            mpatches.Rectangle((0, 0), 1, 1, fc='b')]
    ax.legend(recs, classes, loc=2)

    ax.axis([0, 60, 0, 60])

    plt.show()



def genetic_crossovers_plots():
    wb = load_workbook(filename='genetic_selectparentpartrandom.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    lengths = []
    random = []
    rand_point = []
    halfhalf = []
    lengths.append(sheet['A' + str(2)].value)
    random.append(sheet['D' + str(2)].value)
    rand_point.append(sheet['C' + str(2)].value)
    halfhalf.append(sheet['B' + str(2)].value)
    for i in range(3, 55):
        lengths.append(sheet['A' + str(i)].value)
        halfhalf.append(sheet['B' + str(i)].value + halfhalf[i - 3])
        rand_point.append(sheet['C' + str(i)].value + rand_point[i - 3])
        random.append(sheet['D' + str(i)].value + random[i - 3])

    fig = plt.figure()

    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.set_title('Comparison of Crossover Algorithms')

    ax.set_xlabel('length')
    ax.set_ylabel('cumulative distance from optimal peak sidelobe value')

    ax.scatter(lengths, random, c='r')
    ax.plot(lengths, random, c='r')
    ax.scatter(lengths, rand_point, c='b')
    ax.plot(lengths, rand_point, c='b')
    ax.scatter(lengths, halfhalf, c='g')
    ax.plot(lengths, halfhalf, c='g')

    classes = ['Random', 'Random Point', 'Half Half']
    recs = [mpatches.Rectangle((0, 0), 1, 1, fc='r'), mpatches.Rectangle((0, 0), 1, 1, fc='b'),
            mpatches.Rectangle((0, 0), 1, 1, fc='g')]
    ax.legend(recs, classes, loc=2)

    ax.axis([0, 60, 0, 60])

    plt.show()


def arbitrary_polyphase_comparison_plots():
    random = 53.97306139
    anneal = 33.71477661
    particle_swarm = 39.18536698
    great_deluge = 23.36987512

    objects = ('Random', 'Particle \n Swarm', 'Anneal', 'Great \n Deluge')

    y_pos = np.arange(len(objects))
    performance = [random, particle_swarm, anneal, great_deluge]

    plt.bar(y_pos, performance, align='center', alpha=0.5)
    plt.xticks(y_pos, objects)
    plt.ylabel('Cumulative peak sidelobe, relative to Absolute Barker')
    plt.title('Algorithm Performance for Arbitrary Polyphase Codes')
    plt.show()

def arbitrary_polyphase_comparison_plots_2():
    lengths = []
    random = []
    anneal = []
    particle_swarm = []
    great_deluge = []

    wb = load_workbook(filename='anneal_test_arb_polyphase.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(4, 41):
        lengths.append(sheet['A' + str(i)].value)
        anneal.append(sheet['B' + str(i)].value)

    wb = load_workbook(filename='particle_swarm_test_arb_polyphase.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(4, 41):
        particle_swarm.append(sheet['B' + str(i)].value)

    wb = load_workbook(filename='great_deluge_test_arb_polyphase.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(4, 41):
        great_deluge.append(sheet['B' + str(i)].value)

    wb = load_workbook(filename='random_arbitrary_polyphase.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(2, 39):
        random.append(sheet['D' + str(i)].value- 1)

    fig = plt.figure()

    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.set_title('Algorithm Performance for Arbitrary Polyphase Codes')

    ax.set_xlabel('length')
    ax.set_ylabel('peak sidelobe value')

    ax.scatter(lengths, random, c= 'r')
    ax.plot(lengths, random, c = 'r')
    ax.scatter(lengths, anneal, c = 'b')
    ax.plot(lengths, anneal, c = 'b')
    ax.scatter(lengths, great_deluge, c = 'g')
    ax.plot(lengths, great_deluge, c = 'g')
    ax.scatter(lengths, particle_swarm, c='c')
    ax.plot(lengths, particle_swarm, c='c')

    classes = ['Random', 'Particle Swarm', 'Anneal', 'Great Deluge']
    recs = [mpatches.Rectangle((0,0),1,1,fc='r'), mpatches.Rectangle((0,0),1,1,fc='c'), mpatches.Rectangle((0,0),1,1,fc='b'),
            mpatches.Rectangle((0, 0), 1, 1, fc='g')]
    ax.legend(recs, classes, loc = 2)





    ax.axis([10, 40, 0, 5])

    plt.show()

def new_doppler_codes_hill_climbing_plot():
    lengths = []
    polyphase_3_N_1 = []
    polyphase_3_N_2 = []
    polyphase_3_N_3 = []
    polyphase_3_N_4 = []
    polyphase_3_N_5 = []
    wb = load_workbook(filename='hill_climbing_doppler_polyphase_tests.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(214, 267):
        lengths.append(sheet['D' + str(i)].value)
        polyphase_3_N_1.append(sheet['E' + str(i)].value)

    for i in range(267, 320):
        polyphase_3_N_2.append(sheet['E' + str(i)].value)

    for i in range(320, 373):
        polyphase_3_N_3.append(sheet['E' + str(i)].value)

    for i in range(373, 426):
        polyphase_3_N_4.append(sheet['E' + str(i)].value)

    for i in range(426, 479):
        polyphase_3_N_5.append(sheet['E' + str(i)].value)

    fig = plt.figure()

    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.set_title('Algorithm PSL Values for Doppler, Triphase Codes')

    ax.set_xlabel('length')
    ax.set_ylabel('peak sidelobe value')
    ax.scatter(lengths, polyphase_3_N_1, c='r')
    ax.plot(lengths, polyphase_3_N_1, c='r')
    ax.scatter(lengths, polyphase_3_N_2, c='b')
    ax.plot(lengths, polyphase_3_N_2, c='b')
    ax.scatter(lengths, polyphase_3_N_3, c='g')
    ax.plot(lengths, polyphase_3_N_3, c='g')
    ax.scatter(lengths, polyphase_3_N_4, c='c')
    ax.plot(lengths, polyphase_3_N_4, c='c')
    ax.scatter(lengths, polyphase_3_N_5, c='m')
    ax.plot(lengths, polyphase_3_N_5, c='m')

    classes = ['N=1', 'N=2', 'N=3', 'N=4', 'N=5']
    recs = [mpatches.Rectangle((0, 0), 1, 1, fc='r'), mpatches.Rectangle((0, 0), 1, 1, fc='c'),
            mpatches.Rectangle((0, 0), 1, 1, fc='b'),
            mpatches.Rectangle((0, 0), 1, 1, fc='g'), mpatches.Rectangle((0, 0), 1, 1, fc='m')]
    ax.legend(recs, classes, loc=2)

    ax.axis([0, 55, 0, 12])

    plt.show()

def binary_doppler():
    lengths = []
    polyphase_2_N_2 = []
    polyphase_2_N_3 = []
    polyphase_2_N_4 = []
    polyphase_2_N_5 = []
    wb = load_workbook(filename='hill_climbing_doppler_polyphase_tests.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(2, 55):
        lengths.append(sheet['D' + str(i)].value)
        polyphase_2_N_2.append(sheet['E' + str(i)].value)

    for i in range(55, 108):
        polyphase_2_N_3.append(sheet['E' + str(i)].value)

    for i in range(108, 161):
        polyphase_2_N_4.append(sheet['E' + str(i)].value)

    for i in range(161, 214):
        polyphase_2_N_5.append(sheet['E' + str(i)].value)


    fig = plt.figure()

    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.set_title('Algorithm PSL Values for Doppler, Binary Codes')

    ax.set_xlabel('length')
    ax.set_ylabel('peak sidelobe value')
    ax.scatter(lengths, polyphase_2_N_2, c='b')
    ax.plot(lengths, polyphase_2_N_2, c='b')
    ax.scatter(lengths, polyphase_2_N_3, c='g')
    ax.plot(lengths, polyphase_2_N_3, c='g')
    ax.scatter(lengths, polyphase_2_N_4, c='c')
    ax.plot(lengths, polyphase_2_N_4, c='c')
    ax.scatter(lengths, polyphase_2_N_5, c='m')
    ax.plot(lengths, polyphase_2_N_5, c='m')

    classes = ['N=2', 'N=3', 'N=4', 'N=5']
    recs = [mpatches.Rectangle((0, 0), 1, 1, fc='c'),
            mpatches.Rectangle((0, 0), 1, 1, fc='b'),
            mpatches.Rectangle((0, 0), 1, 1, fc='g'), mpatches.Rectangle((0, 0), 1, 1, fc='m')]
    ax.legend(recs, classes, loc=2)

    ax.axis([0, 55, 0, 12])

    plt.show()

def quad_doppler():
    lengths = []
    polyphase_4_N_1 = []
    polyphase_4_N_2 = []
    polyphase_4_N_3 = []
    polyphase_4_N_4 = []
    polyphase_4_N_5 = []
    wb = load_workbook(filename='hill_climbing_doppler_polyphase_tests.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(479, 532):
        lengths.append(sheet['D' + str(i)].value)
        polyphase_4_N_1.append(sheet['E' + str(i)].value)

    for i in range(532, 585):
        polyphase_4_N_2.append(sheet['E' + str(i)].value)

    for i in range(585, 638):
        polyphase_4_N_3.append(sheet['E' + str(i)].value)

    for i in range(638, 691):
        polyphase_4_N_4.append(sheet['E' + str(i)].value)

    for i in range(691, 744):
        polyphase_4_N_5.append(sheet['E' + str(i)].value)

    fig = plt.figure()

    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.set_title('Algorithm PSL Values for Doppler, Quadphase Codes')

    ax.set_xlabel('length')
    ax.set_ylabel('peak sidelobe value')
    ax.scatter(lengths, polyphase_4_N_1, c='r')
    ax.plot(lengths, polyphase_4_N_1, c='r')
    ax.scatter(lengths, polyphase_4_N_2, c='b')
    ax.plot(lengths, polyphase_4_N_2, c='b')
    ax.scatter(lengths, polyphase_4_N_3, c='g')
    ax.plot(lengths, polyphase_4_N_3, c='g')
    ax.scatter(lengths, polyphase_4_N_4, c='c')
    ax.plot(lengths, polyphase_4_N_4, c='c')
    ax.scatter(lengths, polyphase_4_N_5, c='m')
    ax.plot(lengths, polyphase_4_N_5, c='m')

    classes = ['N=1', 'N=2', 'N=3', 'N=4', 'N=5']
    recs = [mpatches.Rectangle((0, 0), 1, 1, fc='r'), mpatches.Rectangle((0, 0), 1, 1, fc='c'),
            mpatches.Rectangle((0, 0), 1, 1, fc='b'),
            mpatches.Rectangle((0, 0), 1, 1, fc='g'), mpatches.Rectangle((0, 0), 1, 1, fc='m')]
    ax.legend(recs, classes, loc=2)

    ax.axis([0, 55, 0, 12])

    plt.show()

def quint_doppler():
    lengths = []
    polyphase_5_N_1 = []
    polyphase_5_N_2 = []
    polyphase_5_N_3 = []
    polyphase_5_N_4 = []
    polyphase_5_N_5 = []
    wb = load_workbook(filename='hill_climbing_doppler_polyphase_tests.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(744, 797):
        lengths.append(sheet['D' + str(i)].value)
        polyphase_5_N_1.append(sheet['E' + str(i)].value)

    for i in range(797, 850):
        polyphase_5_N_2.append(sheet['E' + str(i)].value)

    for i in range(850, 903):
        polyphase_5_N_3.append(sheet['E' + str(i)].value)

    for i in range(903, 956):
        polyphase_5_N_4.append(sheet['E' + str(i)].value)

    for i in range(956, 1009):
        polyphase_5_N_5.append(sheet['E' + str(i)].value)

    fig = plt.figure()

    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.set_title('Algorithm PSL Values for Doppler, Polyphase Codes with m=5')

    ax.set_xlabel('length')
    ax.set_ylabel('peak sidelobe value')
    ax.scatter(lengths, polyphase_5_N_1, c='r')
    ax.plot(lengths, polyphase_5_N_1, c='r')
    ax.scatter(lengths, polyphase_5_N_2, c='b')
    ax.plot(lengths, polyphase_5_N_2, c='b')
    ax.scatter(lengths, polyphase_5_N_3, c='g')
    ax.plot(lengths, polyphase_5_N_3, c='g')
    ax.scatter(lengths, polyphase_5_N_4, c='c')
    ax.plot(lengths, polyphase_5_N_4, c='c')
    ax.scatter(lengths, polyphase_5_N_5, c='m')
    ax.plot(lengths, polyphase_5_N_5, c='m')

    classes = ['N=1', 'N=2', 'N=3', 'N=4', 'N=5']
    recs = [mpatches.Rectangle((0, 0), 1, 1, fc='r'), mpatches.Rectangle((0, 0), 1, 1, fc='c'),
            mpatches.Rectangle((0, 0), 1, 1, fc='b'),
            mpatches.Rectangle((0, 0), 1, 1, fc='g'), mpatches.Rectangle((0, 0), 1, 1, fc='m')]
    ax.legend(recs, classes, loc=2)

    ax.axis([0, 55, 0, 12])

    plt.show()

quint_doppler()


def binary_codes_cumulative_plot():
    memetic_ts = 6
    memetic_sdls = 12
    anneal_ts = 25
    evolut_reg = 22
    evolut_ls = 24
    hill_climbing_simple = 15
    hill_climbing_sdls = 16
    threshold_accept = 32
    random = 30
    objects = ('Threshold \nAccept', 'Random', 'Anneal', 'Evolutionary',
               'Hill \nClimbing', 'Memetic')

    y_pos = np.arange(len(objects))
    performance = [threshold_accept, random, anneal_ts, evolut_reg, hill_climbing_simple,  memetic_ts]

    plt.bar(y_pos, performance, align = 'center', alpha = 0.5)
    plt.xticks(y_pos, objects)
    plt.ylabel('Number of errors relative to optimal')
    plt.title('Optimization Algorithm Performance for Binary Codes of Lengths 2-55')
    plt.show()

def temperature_cumulative_plot():
    lin_add = 23
    lin_mult = 16
    log_mult = 30
    quad_mult = 20
    quenching = 16
    semiexponential = 17
    objects = ('Mult. \nLogarithmic', 'Additive \nLinear',
               'Mult. \nQuadratic', 'Semi-\nexponential', 'Mult. \nLinear',
               'Quenching')

    y_pos = np.arange(len(objects))
    performance = [log_mult, lin_add, quad_mult, semiexponential, lin_mult, quenching]

    plt.bar(y_pos, performance, align = 'center', alpha = 0.8)
    plt.xticks(y_pos, objects)
    plt.ylabel('Cumulative Rank')
    plt.title('Relative Success of Differing Temperature Schedules')
    plt.show()


def random_code():
    code = [1, 1, 1, -1, -1, -1, -1, 1, 1, 1]
    x_1 = np.linspace(0, math.pi, 200)
    y_1 = np.sin(x_1)
    sin_fns = []
    for i in range(len(code)):
        x = np.linspace(2*math.pi * i + math.pi/2, 2*math.pi * (i+1) + math.pi/2, 200)
        y = -np.sin(x) * code[i]
        sin_fns.append((x, y))

    fig = plt.figure()
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.add_patch(mpatches.Rectangle((math.pi/2, 0), 6*math.pi, 1, fill=True, color='g', alpha = 0.5))
    ax.add_patch(mpatches.Rectangle((math.pi/2 + 6*math.pi, -1), 8 * math.pi, 1, fill=True, color='g', alpha = 0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 14 * math.pi, 0), 6 * math.pi, 1, fill=True, color='g', alpha = 0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 6*math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 14 * math.pi, -1), 0, 2, fill=True, color='#000000'))

    for i in range(len(sin_fns)):
        ax.plot(sin_fns[i][0], sin_fns[i][1], c = 'r')
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)

    ax.axis([0, 70, -5, 5])
    ax.axis('off')

    plt.show()

def autocorrelation_of_that_code():
    autocorrelation = [1, 2, 3, 0, -3, -6, -5, 0, 5, 10, 5, 0, -5, -6, -3, 0, 3, 2, 1]
    x = []
    for i in range(len(autocorrelation)):
        x.append(i)

    fig = plt.figure()
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.scatter(x, np.abs(autocorrelation))
    ax.plot(x, np.abs(autocorrelation))
    ax.axis([-1, 20, -1, 12])
    ax.axis('off')
    plt.show()

def first_alternating_code():
    code = [-1, -1, -1, 1, -1,  1, -1,  1,  1,  1, -1, -1, -1]
    x_1 = np.linspace(0, math.pi, 200)
    y_1 = np.sin(x_1)
    sin_fns = []
    for i in range(len(code)):
        x = np.linspace(2 * math.pi * i + math.pi / 2, 2 * math.pi * (i + 1) + math.pi / 2, 200)
        y = -np.sin(x) * code[i]
        sin_fns.append((x, y))

    fig = plt.figure()
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.add_patch(mpatches.Rectangle((math.pi / 2, -1), 6 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 6 * math.pi, 0), 2 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 8 * math.pi, -1), 2 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 10 * math.pi, 0), 2 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 12 * math.pi, -1), 2 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 14 * math.pi, 0), 6 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 20 * math.pi, -1), 6 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 6 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 8 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 10 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 12 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 14 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 20 * math.pi, -1), 0, 2, fill=True, color='#000000'))

    for i in range(len(sin_fns)):
        ax.plot(sin_fns[i][0], sin_fns[i][1], c='r')
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)

    ax.axis([0, 100, -5, 5])
    ax.axis('off')

    plt.show()
    plt.clf()
    autocorrelation = con_psl_matrix(code)
    x = []
    for i in range(len(autocorrelation)):
        x.append(i)

    fig = plt.figure()
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.scatter(x, np.abs(autocorrelation))
    ax.plot(x, np.abs(autocorrelation))
    ax.axis([-1, 30, -1, 15])
    ax.axis('off')
    plt.show()

def second_alternating_code():
    code = [-1, 1, 1, -1, -1, -1, -1, -1, 1, 1, -1, -1, -1]
    x_1 = np.linspace(0, math.pi, 200)
    y_1 = np.sin(x_1)
    sin_fns = []
    for i in range(len(code)):
        x = np.linspace(2 * math.pi * i + math.pi / 2, 2 * math.pi * (i + 1) + math.pi / 2, 200)
        y = -np.sin(x) * code[i]
        sin_fns.append((x, y))

    fig = plt.figure()
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.add_patch(mpatches.Rectangle((math.pi / 2, -1), 2 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 2 * math.pi, 0), 4 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 6 * math.pi, -1), 10 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 16 * math.pi, 0), 4 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 20 * math.pi, -1), 6 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 2 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 6 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 16 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 20 * math.pi, -1), 0, 2, fill=True, color='#000000'))

    for i in range(len(sin_fns)):
        ax.plot(sin_fns[i][0], sin_fns[i][1], c='r')
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)

    ax.axis([0, 100, -5, 5])
    ax.axis('off')

    plt.show()
    plt.clf()

    autocorrelation = con_psl_matrix(code)
    x = []
    for i in range(len(autocorrelation)):
        x.append(i)

    fig = plt.figure()
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.scatter(x, np.abs(autocorrelation))
    ax.plot(x, np.abs(autocorrelation))
    ax.axis([-1, 30, -1, 15])
    ax.axis('off')
    plt.show()
    plt.clf()

def third_alternating_code():
    code = [1, 1, -1, -1, -1, -1, -1, -1, 1, -1, -1, -1, -1]
    x_1 = np.linspace(0, math.pi, 200)
    y_1 = np.sin(x_1)
    sin_fns = []
    for i in range(len(code)):
        x = np.linspace(2 * math.pi * i + math.pi / 2, 2 * math.pi * (i + 1) + math.pi / 2, 200)
        y = -np.sin(x) * code[i]
        sin_fns.append((x, y))

    fig = plt.figure()
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.add_patch(mpatches.Rectangle((math.pi / 2, 0), 4 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 4 * math.pi, -1), 12 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 16 * math.pi, 0), 2 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 18 * math.pi, -1), 8 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 4 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 16 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 18 * math.pi, -1), 0, 2, fill=True, color='#000000'))

    for i in range(len(sin_fns)):
        ax.plot(sin_fns[i][0], sin_fns[i][1], c='r')
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)

    ax.axis([0, 100, -5, 5])
    ax.axis('off')

    plt.show()
    plt.clf()

    autocorrelation = con_psl_matrix(code)
    x = []
    for i in range(len(autocorrelation)):
        x.append(i)

    fig = plt.figure()
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.scatter(x, np.abs(autocorrelation))
    ax.plot(x, np.abs(autocorrelation))
    ax.axis([-1, 30, -1, 15])
    ax.axis('off')
    plt.show()
    plt.clf()

def fourth_alternating_code():
    code = [-1, 1, 1, -1, -1, 1, -1,  1, -1, -1, 1, -1, -1]
    x_1 = np.linspace(0, math.pi, 200)
    y_1 = np.sin(x_1)
    sin_fns = []
    for i in range(len(code)):
        x = np.linspace(2 * math.pi * i + math.pi / 2, 2 * math.pi * (i + 1) + math.pi / 2, 200)
        y = -np.sin(x) * code[i]
        sin_fns.append((x, y))

    fig = plt.figure()
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.add_patch(mpatches.Rectangle((math.pi / 2, -1), 2 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 2 * math.pi, 0), 4 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 6 * math.pi, -1), 4 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 10 * math.pi, 0), 2 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 12 * math.pi, -1), 2 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 14 * math.pi, 0), 2 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 16 * math.pi, -1), 4 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 20 * math.pi, 0), 2 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 22 * math.pi, -1), 4 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 2 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 6 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 10 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 12 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 14 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 16 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 20 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 22 * math.pi, -1), 0, 2, fill=True, color='#000000'))

    for i in range(len(sin_fns)):
        ax.plot(sin_fns[i][0], sin_fns[i][1], c='r')
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)

    ax.axis([0, 100, -5, 5])
    ax.axis('off')

    plt.show()
    plt.clf()
    autocorrelation = con_psl_matrix(code)
    x = []
    for i in range(len(autocorrelation)):
        x.append(i)

    fig = plt.figure()
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.scatter(x, np.abs(autocorrelation))
    ax.plot(x, np.abs(autocorrelation))
    ax.axis([-1, 30, -1, 15])
    ax.axis('off')
    plt.show()
    plt.clf()

def barker_code():
    code = [-1, -1, -1, -1, -1, 1, 1, -1, -1, 1, -1, 1, -1]
    x_1 = np.linspace(0, math.pi, 200)
    y_1 = np.sin(x_1)
    sin_fns = []
    for i in range(len(code)):
        x = np.linspace(2 * math.pi * i + math.pi / 2, 2 * math.pi * (i + 1) + math.pi / 2, 200)
        y = -np.sin(x) * code[i]
        sin_fns.append((x, y))

    fig = plt.figure()
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.add_patch(mpatches.Rectangle((math.pi / 2, -1), 10 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 10 * math.pi, 0), 4 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 14 * math.pi, -1), 4 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 18 * math.pi, 0), 2* math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 20 * math.pi, -1), 2 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 22 * math.pi, 0), 2 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 24 * math.pi, -1), 2 * math.pi, 1, fill=True, color='g', alpha=0.5))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 10 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 14 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 18 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 20 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 22 * math.pi, -1), 0, 2, fill=True, color='#000000'))
    ax.add_patch(mpatches.Rectangle((math.pi / 2 + 24 * math.pi, -1), 0, 2, fill=True, color='#000000'))

    for i in range(len(sin_fns)):
        ax.plot(sin_fns[i][0], sin_fns[i][1], c='r')
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)

    ax.axis([0, 100, -5, 5])
    ax.axis('off')

    plt.show()
    plt.clf()
    autocorrelation = con_psl_matrix(code)
    x = []
    for i in range(len(autocorrelation)):
        x.append(i)

    fig = plt.figure()
    ax = fig.add_subplot(111)
    fig.subplots_adjust(top=0.85)
    ax.scatter(x, np.abs(autocorrelation))
    ax.plot(x, np.abs(autocorrelation))
    ax.axis([-1, 30, -1, 15])
    ax.axis('off')
    plt.show()
    plt.clf()
